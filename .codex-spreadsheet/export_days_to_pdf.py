from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle


SOURCE_PATH = Path(os.environ["XLSX_PATH"])
OUTPUT_DIR = Path(os.environ["OUTPUT_DIR"])

FONT_REGULAR = "MSYH"
FONT_BOLD = "MSYH-Bold"


def register_fonts() -> None:
    pdfmetrics.registerFont(TTFont(FONT_REGULAR, r"C:\Windows\Fonts\msyh.ttc"))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, r"C:\Windows\Fonts\msyhbd.ttc"))


def normalize(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def build_table_data(ws):
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "header",
        parent=styles["Normal"],
        fontName=FONT_BOLD,
        fontSize=10,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    body_style = ParagraphStyle(
        "body",
        parent=styles["Normal"],
        fontName=FONT_REGULAR,
        fontSize=8.5,
        leading=11,
        wordWrap="CJK",
    )

    data = []
    for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
        rendered = []
        for value in row:
            text = normalize(value)
            style = header_style if len(data) == 0 else body_style
            rendered.append(Paragraph(text, style))
        data.append(rendered)
    return data


def export_sheet(ws, output_path: Path) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    page_width, page_height = landscape(A4)
    left_margin = 8 * mm
    right_margin = 8 * mm
    top_margin = 8 * mm
    bottom_margin = 8 * mm
    usable_width = page_width - left_margin - right_margin

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
        title=ws.title,
    )

    data = build_table_data(ws)
    width_weights = [8, 16, 16, 24, 42]
    total = sum(width_weights)
    col_widths = [usable_width * weight / total for weight in width_weights]

    table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d3e0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f9fc")]),
            ]
        )
    )
    doc.build([table])


def main() -> None:
    register_fonts()
    wb = load_workbook(SOURCE_PATH, data_only=False)

    for ws in wb.worksheets:
        if not ws.title.startswith("Day"):
            continue
        output_path = OUTPUT_DIR / f"{ws.title}.pdf"
        export_sheet(ws, output_path)
        print(output_path)


if __name__ == "__main__":
    main()
