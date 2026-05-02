from __future__ import annotations

import os
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


SOURCE_PATH = Path(os.environ["XLSX_PATH"])
HEADERS = ["序号", "词汇", "音标", "核心释义", "高频搭配"]


def copy_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def style_sheet_for_print(ws, last_row: int) -> None:
    ws.print_area = f"$A$1:$E${last_row}"
    ws.print_title_rows = "$1:$1"
    ws.sheet_view.view = "pageLayout"
    ws.print_options.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.12
    ws.page_margins.footer = 0.12
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = "A2"

    # Show only the five retained columns in the sheet view.
    ws.column_dimensions.group("F", "XFD", hidden=True)

    widths = {
        "A": 8,
        "B": 16,
        "C": 16,
        "D": 24,
        "E": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, last_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.alignment = copy(cell.alignment)
            cell.alignment = cell.alignment.copy(wrapText=True, vertical="top")

    ws.row_dimensions[1].height = 24


def build_day1(source_ws, target_ws) -> int:
    last_row = 1
    for row in range(1, source_ws.max_row + 1):
        if any(source_ws.cell(row, col).value not in (None, "") for col in range(1, 6)):
            last_row = row

    for row in range(1, last_row + 1):
        for col in range(1, 6):
            src = source_ws.cell(row, col)
            dst = target_ws.cell(row, col, src.value)
            copy_style(src, dst)

    return last_row


def build_dayn(source_ws, target_ws) -> int:
    source_header_row = 9
    source_data_start = 10
    target_ws.append(HEADERS)

    for col in range(1, 6):
        copy_style(source_ws.cell(source_header_row, min(col, 4) if col < 5 else 11), target_ws.cell(1, col))

    target_row = 2
    for row in range(source_data_start, source_ws.max_row + 1):
        serial = source_ws.cell(row, 1).value
        word = source_ws.cell(row, 2).value
        phonetic = source_ws.cell(row, 3).value
        meaning = source_ws.cell(row, 4).value
        collocation = source_ws.cell(row, 11).value

        if all(value in (None, "") for value in (serial, word, phonetic, meaning, collocation)):
            continue

        values = [serial, word, phonetic, meaning, collocation]
        for col, value in enumerate(values, start=1):
            src_col = col if col < 5 else 11
            src = source_ws.cell(row, src_col)
            dst = target_ws.cell(target_row, col, value)
            copy_style(src, dst)
        target_row += 1

    return target_row - 1


def main() -> None:
    source_wb = load_workbook(SOURCE_PATH)
    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    configured = []
    for index in range(1, 15):
        name = f"Day{index}"
        if name not in source_wb.sheetnames:
            continue
        source_ws = source_wb[name]
        target_ws = output_wb.create_sheet(title=name)

        if name == "Day1":
            last_row = build_day1(source_ws, target_ws)
        else:
            last_row = build_dayn(source_ws, target_ws)

        style_sheet_for_print(target_ws, last_row)
        configured.append((name, last_row))

    output_path = SOURCE_PATH.with_name(
        f"{SOURCE_PATH.stem}-仅保留词表5列-A4打印优化版.xlsx"
    )
    output_wb.save(output_path)

    print(output_path)
    for name, last_row in configured:
        print(f"{name}\tA1:E{last_row}")


if __name__ == "__main__":
    main()
